"""Excel ingestion — read_excel.

Finance and business data live in `.xlsx`. This surfaces a workbook's structure
(sheets, then a sheet's columns / row count / sample) so the model can reason
about it. Like read_csv, it *surfaces* data; exact computation stays in code
(calculate / aggregate over the values).
"""

from __future__ import annotations

import zipfile
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from halia.permissions.guard import check_readable

_MAX_ROWS = 100_000
_DEFAULT_SAMPLE = 10


def extract_excel_text(
    path: Path, sample_rows: int = _DEFAULT_SAMPLE, max_chars: int = 10_000
) -> str:
    """Render a workbook as text (each sheet's columns + a sample) — for teach/reference.

    No permission-floor check (caller owns the file). Raises InvalidFileException/OSError/
    BadZipFile on a malformed file.
    """
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        blocks: list[str] = []
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            header: list[str] | None = None
            sample: list[list[str]] = []
            count = 0
            for i, row in enumerate(worksheet.iter_rows(values_only=True)):
                cells = ["" if value is None else str(value) for value in row]
                if i == 0:
                    header = cells
                    continue
                count += 1
                if len(sample) < sample_rows:
                    sample.append(cells)
                if count >= _MAX_ROWS:
                    break
            lines = [f"=== sheet: {sheet} ==="]
            if header is None:
                lines.append("(empty)")
            else:
                lines.append(f"columns ({len(header)}): {', '.join(header)}")
                lines.append(f"data rows: {count}")
                lines.append(" | ".join(header))
                lines.extend(" | ".join(cells) for cells in sample)
            blocks.append("\n".join(lines))
    finally:
        workbook.close()

    text = "\n\n".join(blocks) or "(empty workbook)"
    if len(text) > max_chars:
        text = text[:max_chars] + "\n… [truncated]"
    return text


class ReadExcel:
    name = "read_excel"
    description = (
        "Read an Excel (.xlsx) file. With no 'sheet', lists the sheet names; with a "
        "sheet, returns its columns, row count, and a sample of rows."
    )
    dangerous = False
    untrusted = False
    parameters: dict[str, Any] = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "path": {"type": "string", "description": "Path to the .xlsx file."},
            "sheet": {"type": "string", "description": "Sheet name (omit to list sheets)."},
            "sample_rows": {
                "type": "integer",
                "description": "How many data rows to sample (default 10).",
            },
        },
        "required": ["path"],
    }

    def run(self, args: dict[str, Any]) -> str:
        raw = args.get("path")
        if not isinstance(raw, str) or not raw.strip():
            return "error: 'path' is required and must be a non-empty string"
        path = Path(raw).expanduser()
        check_readable(path)
        if not path.is_file():
            return f"error: not a file: {path}"

        sample_n = args.get("sample_rows", _DEFAULT_SAMPLE)
        if not isinstance(sample_n, int) or sample_n <= 0:
            sample_n = _DEFAULT_SAMPLE

        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except (InvalidFileException, OSError, zipfile.BadZipFile) as exc:
            return f"error reading Excel: {exc}"

        try:
            sheet = args.get("sheet")
            if not isinstance(sheet, str) or not sheet.strip():
                return "sheets: " + ", ".join(workbook.sheetnames)
            if sheet not in workbook.sheetnames:
                return f"error: sheet '{sheet}' not found. Sheets: {', '.join(workbook.sheetnames)}"

            worksheet = workbook[sheet]
            header: list[str] | None = None
            sample: list[list[str]] = []
            count = 0
            truncated = False
            for i, row in enumerate(worksheet.iter_rows(values_only=True)):
                cells = ["" if value is None else str(value) for value in row]
                if i == 0:
                    header = cells
                    continue
                count += 1
                if len(sample) < sample_n:
                    sample.append(cells)
                if count >= _MAX_ROWS:
                    truncated = True
                    break
        finally:
            workbook.close()

        if header is None:
            return f"(sheet '{sheet}' is empty)"

        lines = [
            f"sheet: {sheet}",
            f"columns ({len(header)}): {', '.join(header)}",
            f"data rows: {count}{'+' if truncated else ''}",
            "sample:",
            " | ".join(header),
        ]
        lines.extend(" | ".join(cells) for cells in sample)
        return "\n".join(lines)
