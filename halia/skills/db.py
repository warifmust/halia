"""Structured query skill — read-only SQL over a SQLite database.

Business data lives in databases; this lets the agent query it directly (agentic
search over structured data), with the DB engine doing the computation. Two
layers of safety: the query must be a SELECT/WITH, AND the connection is opened
read-only (`mode=ro`), so a write is rejected by SQLite itself even if the text
check were bypassed.
"""

from __future__ import annotations

import csv
import re
import sqlite3
from pathlib import Path
from typing import Any

from halia.permissions.guard import check_readable

_MAX_ROWS = 100
_LOAD_MAX = 100_000  # cap rows loaded per file


class QueryDb:
    name = "query_db"
    description = (
        "Run a READ-ONLY SQL query (SELECT / WITH … SELECT only) against a SQLite "
        "database file and return the rows. Use SQL aggregates (SUM, COUNT, …) for "
        "exact computation over the data."
    )
    dangerous = False
    untrusted = False
    parameters: dict[str, Any] = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "path": {"type": "string", "description": "Path to the SQLite database file."},
            "query": {"type": "string", "description": "A read-only SQL SELECT query."},
        },
        "required": ["path", "query"],
    }

    def run(self, args: dict[str, Any]) -> str:
        raw_path = args.get("path")
        if not isinstance(raw_path, str) or not raw_path.strip():
            return "error: 'path' is required and must be a non-empty string"
        query = args.get("query")
        if not isinstance(query, str) or not query.strip():
            return "error: 'query' is required and must be a non-empty string"

        statement = query.strip().rstrip(";").strip()
        if not statement.lower().startswith(("select", "with")):
            return "error: only read-only SELECT (or WITH … SELECT) queries are allowed"

        path = Path(raw_path).expanduser()
        check_readable(path)
        if not path.is_file():
            return f"error: not a file: {path}"

        try:
            conn = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
            try:
                cursor = conn.execute(statement)  # single statement; RO rejects writes
                columns = [d[0] for d in cursor.description] if cursor.description else []
                rows = cursor.fetchmany(_MAX_ROWS + 1)
            finally:
                conn.close()
        except sqlite3.Error as exc:
            return f"error running query: {exc}"

        if not columns:
            return "(query returned no columns)"

        truncated = len(rows) > _MAX_ROWS
        lines = [" | ".join(columns)]
        lines.extend(" | ".join(str(value) for value in row) for row in rows[:_MAX_ROWS])
        if truncated:
            lines.append(f"… [truncated at {_MAX_ROWS} rows]")
        return "\n".join(lines)


def _table_name(stem: str, taken: set[str]) -> str:
    name = re.sub(r"\W", "_", stem).strip("_").lower() or "t"
    if name[0].isdigit():
        name = f"t_{name}"
    base, i = name, 2
    while name in taken:
        name, i = f"{base}_{i}", i + 1
    taken.add(name)
    return name


def _read_rows(path: Path) -> list[list[str]]:
    """Read a CSV or Excel file into a list of string rows (header first)."""
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        wb = load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        rows: list[list[str]] = []
        for r in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else str(c) for c in r])
            if len(rows) > _LOAD_MAX:
                break
        wb.close()
        return rows
    with path.open(newline="", encoding="utf-8", errors="replace") as handle:
        out: list[list[str]] = []
        for row in csv.reader(handle):
            out.append(row)
            if len(out) > _LOAD_MAX:
                break
        return out


def _column_type(values: list[str]) -> str:
    seen = [v.strip() for v in values if v.strip() != ""]
    if not seen:
        return "TEXT"
    cleaned = [v.lstrip("$").replace(",", "") for v in seen]
    if all(re.fullmatch(r"-?\d+", c) for c in cleaned):
        return "INTEGER"
    if all(re.fullmatch(r"-?\d*\.?\d+", c) for c in cleaned):
        return "REAL"
    return "TEXT"


def _convert(value: str, sql_type: str) -> Any:
    if value.strip() == "":
        return None
    if sql_type == "INTEGER":
        return int(value.strip().lstrip("$").replace(",", ""))
    if sql_type == "REAL":
        return float(value.strip().lstrip("$").replace(",", ""))
    return value


def _load_table(conn: sqlite3.Connection, table: str, rows: list[list[str]]) -> list[str]:
    """Create + populate a typed table from string rows; return its column names."""
    header = rows[0] if rows else []
    body = rows[1:]
    ncols = len(header)
    types = [_column_type([r[c] if c < len(r) else "" for r in body]) for c in range(ncols)]
    cols_ddl = ", ".join(f'"{header[c]}" {types[c]}' for c in range(ncols))
    conn.execute(f'CREATE TABLE "{table}" ({cols_ddl})')
    placeholders = ", ".join("?" * ncols)
    conn.executemany(
        f'INSERT INTO "{table}" VALUES ({placeholders})',
        [[_convert(r[c] if c < len(r) else "", types[c]) for c in range(ncols)] for r in body],
    )
    return list(header)


class QueryData:
    name = "query_data"
    description = (
        "Run read-only SQL (SELECT / WITH … SELECT) over one or more CSV/Excel files. Each "
        "file is loaded into an in-memory table named after its filename without the "
        "extension (non-letters become '_'). Gives you JOIN, WHERE, ORDER BY, GROUP BY, "
        "subqueries over flat files. Numeric columns are typed so sorting and aggregation "
        "work; quote column names with spaces as \"like this\"."
    )
    dangerous = False
    untrusted = False
    parameters: dict[str, Any] = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "files": {
                "type": "array",
                "items": {"type": "string"},
                "description": "Paths to CSV/Excel files to load as tables.",
            },
            "query": {"type": "string", "description": "A read-only SQL SELECT query."},
        },
        "required": ["files", "query"],
    }

    def run(self, args: dict[str, Any]) -> str:
        files = args.get("files")
        query = args.get("query")
        if not isinstance(files, list) or not files or not all(isinstance(f, str) for f in files):
            return "error: 'files' must be a non-empty array of file paths"
        if not isinstance(query, str) or not query.strip():
            return "error: 'query' is required"
        statement = query.strip().rstrip(";").strip()
        if not statement.lower().startswith(("select", "with")):
            return "error: only read-only SELECT (or WITH … SELECT) queries are allowed"

        conn = sqlite3.connect(":memory:")
        schema: dict[str, list[str]] = {}
        try:
            taken: set[str] = set()
            for raw in files:
                path = Path(raw).expanduser()
                check_readable(path)
                if not path.is_file():
                    return f"error: not a file: {path}"
                if path.suffix.lower() not in (".csv", ".xlsx", ".xlsm"):
                    return f"error: unsupported file type '{path.suffix}' (use CSV or Excel)"
                rows = _read_rows(path)
                if not rows:
                    return f"error: {path.name} is empty"
                table = _table_name(path.stem, taken)
                schema[table] = _load_table(conn, table, rows)

            try:
                cursor = conn.execute(statement)
                columns = [d[0] for d in cursor.description] if cursor.description else []
                result = cursor.fetchmany(_MAX_ROWS + 1)
            except sqlite3.Error as exc:
                tables = "; ".join(f"{t}({', '.join(cols)})" for t, cols in schema.items())
                return f"error running query: {exc}\nAvailable tables: {tables}"
        finally:
            conn.close()

        if not columns:
            return "(query returned no columns)"
        truncated = len(result) > _MAX_ROWS
        lines = [" | ".join(columns)]
        for row in result[:_MAX_ROWS]:
            lines.append(" | ".join("" if v is None else str(v) for v in row))
        if truncated:
            lines.append(f"… [truncated at {_MAX_ROWS} rows]")
        return "\n".join(lines)
