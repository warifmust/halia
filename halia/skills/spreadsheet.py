"""Spreadsheet output — write grounded tabular data to a real .xlsx.

Unlike the document renders (PDF/PPTX/DOCX, whose master is markdown), a spreadsheet's
master is TABULAR DATA. So this takes rows of values and writes an Excel file with
numbers stored AS numbers (so the analyst can SUM / pivot / chart), bold headers, and a
frozen header row. Same trust division as slides: halia hands over grounded data; the
finance person takes the wheel in Excel. `.xlsx` also imports straight into Google Sheets.

Uses the existing openpyxl dependency (no new dep). v1 = values, not formulas.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from halia.permissions.guard import PermissionDenied, check_writable

_INT = re.compile(r"-?\d+")
_FLOAT = re.compile(r"-?\d*\.\d+")
_MAX_WIDTH = 50


def _coerce(value: Any) -> Any:
    """Store numbers as numbers (so Excel can compute); leave everything else as text."""
    if isinstance(value, bool) or isinstance(value, (int, float)):
        return value
    if not isinstance(value, str):
        return str(value)
    text = value.strip()
    if not text:
        return ""
    cleaned = text.lstrip("$").replace(",", "")  # tolerate "$1,200.50"
    if _INT.fullmatch(cleaned):
        return int(cleaned)
    if _FLOAT.fullmatch(cleaned):
        return float(cleaned)
    return text


def _write_sheet(ws: Any, rows: list[list[Any]]) -> None:
    for r, row in enumerate(rows, start=1):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=_coerce(value))
            if r == 1:
                cell.font = Font(bold=True)
    if rows:
        ws.freeze_panes = "A2"  # keep the header visible when scrolling
        ncols = max(len(row) for row in rows)
        for c in range(1, ncols + 1):
            width = max((len(str(row[c - 1])) for row in rows if c - 1 < len(row)), default=8)
            ws.column_dimensions[get_column_letter(c)].width = min(max(width + 2, 8), _MAX_WIDTH)


def _valid_rows(rows: Any) -> bool:
    return isinstance(rows, list) and len(rows) > 0 and all(isinstance(r, list) for r in rows)


def _check_format_match(headers: list[str]) -> str | None:
    """Check if the header row matches any taught reference format spec.

    Returns a warning string if there's a mismatch, None if headers match or no
    reference is found. The warning tells the model to inform the user.
    """
    from halia.references import list_ref_files

    refs = list_ref_files()
    if not refs:
        return None

    headers_lower = [h.strip().lower() for h in headers]
    for ref in refs:
        if not ref.description:
            continue
        # Extract column names from the format spec description
        # Look for patterns like "Headers: A, B, C" or "columns: A, B, C"
        desc = ref.description.lower()
        for prefix in ("headers:", "columns:", "format:", "fields:"):
            if prefix in desc:
                idx = desc.index(prefix) + len(prefix)
                chunk = desc[idx:idx + 500].split("\n")[0].strip()
                # Strip parenthesized type info before splitting
                # e.g. "test id (string, pattern tc-xxx)" → "test id"
                import re as _re
                chunk = _re.sub(r"\([^)]*\)", "", chunk)
                # Extract comma-separated or pipe-separated column names
                if "," in chunk:
                    spec_cols = [c.strip().strip('"').strip("'")
                                 for c in chunk.split(",")]
                elif "|" in chunk:
                    spec_cols = [c.strip().strip('"').strip("'")
                                 for c in chunk.split("|")]
                else:
                    continue
                spec_cols = [c for c in spec_cols if c and len(c) > 1]
                if not spec_cols:
                    continue
                spec_lower = [c.lower() for c in spec_cols]
                # Check if headers roughly match the spec
                if len(headers_lower) != len(spec_lower):
                    return (
                        f"Format note: reference '{ref.filename}' has "
                        f"{len(spec_lower)} columns ({', '.join(spec_cols[:5])}...) "
                        f"but your output has {len(headers_lower)} columns. "
                        f"Ensure the format matches what was taught."
                    )
                # Check column names (fuzzy — check if most match)
                matches = sum(
                    1 for h, s in zip(headers_lower, spec_lower, strict=True)
                    if h in s or s in h
                )
                if matches < len(spec_lower) * 0.5:
                    return (
                        f"Format note: reference '{ref.filename}' expects "
                        f"columns like {', '.join(spec_cols[:5])}... but your "
                        f"output has {', '.join(headers[:5])}... "
                        f"Ensure the format matches what was taught."
                    )
    return None


class MakeExcel:
    name = "make_excel"
    description = (
        "Write tabular data to an Excel .xlsx file (numbers stored as real numbers so they "
        "can be summed/pivoted; bold header row). Pass `rows` (array of arrays, first row = "
        "headers) for one sheet, or `sheets` (array of {name, rows}) for several. Opens in "
        "Excel and Google Sheets."
    )
    dangerous = True
    untrusted = False  # writes a file
    parameters: dict[str, Any] = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "path": {"type": "string", "description": "Output .xlsx file path."},
            "rows": {
                "type": "array",
                "items": {"type": "array"},
                "description": "Rows of values; first row is the header. (Single sheet.)",
            },
            "sheet_name": {"type": "string", "description": "Sheet name for `rows`."},
            "sheets": {
                "type": "array",
                "items": {"type": "object"},
                "description": "Multiple sheets: each {name, rows}.",
            },
        },
        "required": ["path"],
    }

    def run(self, args: dict[str, Any]) -> str:
        path = args.get("path")
        if not isinstance(path, str) or not path.strip():
            return "error: 'path' is required"

        sheets = args.get("sheets")
        rows = args.get("rows")
        plan: list[tuple[str, Any]] = []
        if isinstance(sheets, list) and sheets:
            for i, sheet in enumerate(sheets):
                if not isinstance(sheet, dict) or not _valid_rows(sheet.get("rows")):
                    return f"error: sheet {i} must be an object with a non-empty 'rows' array"
                plan.append((str(sheet.get("name") or f"Sheet{i + 1}"), sheet["rows"]))
        elif _valid_rows(rows):
            plan.append((str(args.get("sheet_name") or "Sheet1"), rows))
        else:
            return "error: provide 'rows' (array of arrays) or 'sheets' (array of {name, rows})"

        target = Path(path).expanduser()
        try:
            check_writable(target)
        except PermissionDenied as exc:
            return f"blocked: {exc}"

        # Format validation: check headers against taught reference specs.
        warnings: list[str] = []
        for _, sheet_rows in plan:
            if sheet_rows and isinstance(sheet_rows[0], list) and sheet_rows[0]:
                fmt_warn = _check_format_match(
                    [str(h) for h in sheet_rows[0]]
                )
                if fmt_warn:
                    warnings.append(fmt_warn)

        wb = Workbook()
        wb.remove(wb.active)  # drop the default sheet; we add our own
        for name, sheet_rows in plan:
            _write_sheet(wb.create_sheet(title=name[:31]), sheet_rows)
        try:
            wb.save(str(target))
        except OSError as exc:
            return f"error writing {path}: {exc}"
        total = sum(len(r) for _, r in plan)
        result = f"wrote {len(plan)} sheet(s), {total} rows to {target}"
        if warnings:
            result += "\n" + "\n".join(warnings)
        return result
