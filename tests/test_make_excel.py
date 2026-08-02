"""Tests for make_excel (tabular data -> .xlsx)."""

from typing import Any

from openpyxl import load_workbook

from halia.skills.spreadsheet import MakeExcel, _coerce


def test_coerce_numbers_and_ids() -> None:
    assert _coerce("$1,200.50") == 1200.5  # currency/thousands -> float
    assert _coerce("90") == 90 and isinstance(_coerce("90"), int)
    assert _coerce("CHK-1042") == "CHK-1042"  # id stays text
    assert _coerce("42%") == "42%"  # not purely numeric -> text
    assert _coerce(78.3) == 78.3


def test_writes_typed_xlsx_with_bold_header(tmp_path: Any) -> None:
    out = tmp_path / "grades.xlsx"
    rows = [["Student", "Average"], ["Aisha", "78.3"], ["Chong", "90"]]
    result = MakeExcel().run({"path": str(out), "rows": rows, "sheet_name": "Grades"})
    assert "wrote 1 sheet(s), 3 rows" in result
    ws = load_workbook(out)["Grades"]
    assert ws["A1"].font.bold is True
    assert ws["B2"].value == 78.3 and isinstance(ws["B2"].value, float)  # numeric -> summable
    assert ws.freeze_panes == "A2"


def test_multiple_sheets(tmp_path: Any) -> None:
    out = tmp_path / "book.xlsx"
    r = MakeExcel().run(
        {
            "path": str(out),
            "sheets": [
                {"name": "Ledger", "rows": [["id", "amt"], ["A", "10"]]},
                {"name": "Bank", "rows": [["id", "amt"], ["A", "12"]]},
            ],
        }
    )
    assert "wrote 2 sheet(s)" in r
    assert load_workbook(out).sheetnames == ["Ledger", "Bank"]


def test_requires_rows_or_sheets(tmp_path: Any) -> None:
    assert "rows" in MakeExcel().run({"path": str(tmp_path / "x.xlsx")})
    assert "rows" in MakeExcel().run({"path": str(tmp_path / "x.xlsx"), "rows": "not a list"})


def test_bad_sheet_reported(tmp_path: Any) -> None:
    r = MakeExcel().run({"path": str(tmp_path / "x.xlsx"), "sheets": [{"name": "X"}]})
    assert "must be an object with a non-empty 'rows'" in r


def test_make_excel_honors_permission_floor(tmp_path: Any) -> None:
    blocked = tmp_path / "credentials.xlsx"
    r = MakeExcel().run({"path": str(blocked), "rows": [["a"], ["1"]]})
    assert r.startswith("blocked:")
    assert not blocked.exists()


def test_make_excel_is_dangerous_and_wired() -> None:
    from halia.presets import get_preset
    from halia.skills import available_skills

    assert MakeExcel().dangerous is True
    assert "make_excel" in available_skills()
    assert "make_excel" in get_preset("finance").skills
